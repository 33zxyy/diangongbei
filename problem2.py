from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from itertools import product
from pathlib import Path
import math
import re
import sys
import zipfile
from typing import Any

import numpy as np
import pandas as pd
try:
    import matplotlib.pyplot as plt
    HAS_MPL = True
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False
except Exception:
    HAS_MPL = False

BUDGET_YUAN = 1_200_000
SERVICE_RADIUS_M = 1000
DEFAULT_WEIGHTS = {'distance': 0.2, 'response': 0.3, 'price': 0.5}


def clean_text(x: Any) -> str:
    if pd.isna(x):
        return ''
    s = str(x).strip()
    for a, b in [('\n', ''), ('\r', ''), (' ', ''), ('\u3000', ''), ('：', ':'), ('（', '('), ('）', ')'), ('－', '-'), ('—', '-')]:
        s = s.replace(a, b)
    return s


def parse_number(x: Any, default=np.nan) -> float:
    if pd.isna(x):
        return default
    if isinstance(x, (int, float, np.integer, np.floating)):
        return float(x)
    s = str(x)
    m = re.search(r'-?\d+(?:\.\d+)?', s)
    if not m:
        return default
    val = float(m.group())
    if '%' in s or '％' in s:
        val /= 100
    return val


HEADER_KEYWORDS = [
    '小区', '社区', '年份', '年度', '老人', '总数', '需求', '实际', '合计',
    '服务', '项目', '规模', '建设', '成本', '固定', '运营', '管理', '能力',
    '距离', '矩阵', '营收', '收入', '支出', '单价', '满意', '权重'
]


def promote_header_if_needed(df: pd.DataFrame, sheet_name: str = '') -> pd.DataFrame:
    """处理 Excel 顶部有标题行、导致列名变成 Unnamed:1/2/3 的情况。

    例如附件3被读成：
    ['服务站建设与运营成本', 'Unnamed:1', 'Unnamed:2', 'Unnamed:3']，
    真正的表头其实在下一行：['规模', '建设成本', '日固定成本', '日服务能力']。
    """
    cdf = df.copy()
    cdf.columns = [clean_text(c) for c in cdf.columns]

    cols_txt = '|'.join(cdf.columns)
    unnamed_cnt = sum((c == '') or c.lower().startswith('unnamed') for c in cdf.columns)
    col_hits = sum(1 for k in HEADER_KEYWORDS if k in cols_txt)

    # 如果当前列名已经像正常表头，就不动。
    if unnamed_cnt == 0 and col_hits >= 2:
        return cdf

    best_idx = None
    best_score = -1
    # 只在前若干行找真正表头，避免把数据行误当成表头。
    for idx in range(min(12, len(cdf))):
        vals = [clean_text(v) for v in cdf.iloc[idx].tolist()]
        nonempty = sum(v != '' and v.lower() != 'nan' for v in vals)
        row_txt = '|'.join(vals)
        hits = sum(1 for k in HEADER_KEYWORDS if k in row_txt)
        # 表头通常包含多个关键词，且非空单元格不止一个。
        score = hits * 10 + nonempty
        if hits >= 2 and nonempty >= 2 and score > best_score:
            best_idx = idx
            best_score = score

    if best_idx is None:
        return cdf

    raw_header = [clean_text(v) for v in cdf.iloc[best_idx].tolist()]
    new_header: list[str] = []
    seen: dict[str, int] = {}
    for j, h in enumerate(raw_header):
        if h == '' or h.lower() == 'nan':
            h = f'列{j+1}'
        if h in seen:
            seen[h] += 1
            h = f'{h}_{seen[h]}'
        else:
            seen[h] = 1
        new_header.append(h)

    out = cdf.iloc[best_idx + 1:].copy()
    out.columns = new_header
    out = out.dropna(how='all').reset_index(drop=True)
    print(f'INFO: sheet「{sheet_name}」检测到标题行/空列名，已将第 {best_idx + 2} 行提升为表头：{new_header}')
    return out



def parse_satisfaction_weights(a5s: dict[str, pd.DataFrame]) -> dict[str, float]:
    """从附件5“满意度评分规则”中解析综合满意度权重。

    只解析综合公式中的权重，例如：
        S = 0.2*S1 + 0.3*S2 + 0.5*S3

    注意：附件5后续评分规则中还会出现 S1=1.00、S2=1.00、S3=1.00，
    这些是单项满意度取值，不是综合权重，不能拿来归一化。
    """
    def norm_formula_text(x: Any) -> str:
        t = clean_text(x)
        t = (t.replace('＝', '=')
               .replace('＊', '*')
               .replace('×', '*')
               .replace('x', '*')
               .replace('X', '*')
               .replace('Ｓ', 'S')
               .replace('s', 'S'))
        # 兼容“0. 2”或“0。2”这类从 Excel 读出来的异常写法。
        t = t.replace('。', '.')
        t = re.sub(r'(\d)\.\s+(\d)', r'\1.\2', t)
        return t

    candidates: list[str] = []
    for _, df in a5s.items():
        for c in df.columns:
            candidates.append(norm_formula_text(c))
        for v in df.astype(str).values.ravel():
            candidates.append(norm_formula_text(v))

    # 优先只看包含完整综合公式的单元格/列名，避免误读后面的 S1=1.00 等评分规则。
    formula_candidates = [
        c for c in candidates
        if all(term in c.upper() for term in ['S1', 'S2', 'S3'])
        and ('S=' in c.upper() or '记为S' in c or '共同决定' in c or '核心因素' in c)
    ]
    # 如果公式被拆散，退一步把所有文本合在一起，但仍然只解析 S1/S2/S3 前面的系数。
    if not formula_candidates:
        formula_candidates = [''.join(candidates)]

    def coef_before(term: str, text: str) -> float | None:
        # 只接受 0.2*S1、0.2S1、20*S1 这种“系数在 S1/S2/S3 前”的综合公式写法。
        # 不接受 S1=1.00，因为那是单项评分规则，不是权重。
        m = re.search(r'(?<![\d.])(\d+(?:\.\d+)?)\s*\*?\s*' + term + r'(?!\d)', text, flags=re.IGNORECASE)
        return float(m.group(1)) if m else None

    for cand in formula_candidates:
        text = cand.upper()
        w1 = coef_before('S1', text)
        w2 = coef_before('S2', text)
        w3 = coef_before('S3', text)
        if w1 is None or w2 is None or w3 is None:
            continue
        ws = np.array([w1, w2, w3], dtype=float)
        if not np.all(np.isfinite(ws)) or ws.sum() <= 0:
            continue
        # 如果表里写 20、30、50，则归一化；若写 0.2、0.3、0.5，保持原值。
        if ws.max() > 1:
            ws = ws / ws.sum()
        # 如果解析结果不是 1 附近，但又明确是公式，仍保留原系数；通常应为 0.2+0.3+0.5=1。
        weights = {'distance': float(ws[0]), 'response': float(ws[1]), 'price': float(ws[2])}
        print('INFO: 已从附件5综合公式解析满意度权重:', weights)
        return weights

    print('WARNING: 未从附件5综合公式解析到 0.2*S1+0.3*S2+0.5*S3，使用默认权重', DEFAULT_WEIGHTS)
    return DEFAULT_WEIGHTS.copy()

def read_workbook_sheets(path: Path) -> dict[str, pd.DataFrame]:
    if path.name.startswith('~$'):
        return {}
    sheets = pd.read_excel(path, sheet_name=None)
    cleaned = {}
    for name, df in sheets.items():
        cdf = promote_header_if_needed(df, sheet_name=name)
        cdf.columns = [clean_text(c) for c in cdf.columns]
        cleaned[name] = cdf
    return cleaned


def choose_sheet_by_keywords(sheets: dict[str, pd.DataFrame], keywords: list[str]) -> tuple[str, pd.DataFrame]:
    """按关键词选择工作表。

    v2 的问题是只在 sheet 名称和前几行数据里找关键词，
    没有搜索列名。附件3的 sheet 名是“服务站建设与运营成本”，
    真正的关键词“规模”在列名“站点规模”里，因此会误报找不到。
    """
    kws = [clean_text(k).lower() for k in keywords]

    # 1) 先匹配 sheet 名称
    for n, df in sheets.items():
        nn = clean_text(n).lower()
        if all(k in nn for k in kws):
            return n, df

    # 2) 再匹配“列名 + 前10行数据”
    for n, df in sheets.items():
        coltxt = '|'.join(clean_text(c).lower() for c in df.columns)
        datatxt = '|'.join(clean_text(v).lower() for v in df.head(10).astype(str).values.ravel())
        fulltxt = coltxt + '|' + datatxt
        if all(k in fulltxt for k in kws):
            return n, df

    # 3) 如果整个文件只有一个 sheet，直接使用它，避免附件3这类单表文件误停。
    if len(sheets) == 1:
        n, df = next(iter(sheets.items()))
        print(f'WARNING: 未严格匹配关键词 {keywords}，但文件只有一个sheet，已使用：{n}')
        return n, df

    print('无法匹配sheet，关键词:', keywords)
    for n, df in sheets.items():
        print(f'- {n}: cols={list(df.columns)}')
        print(df.head(3))
    raise KeyError(f'无法找到包含关键词 {keywords} 的sheet')


def pick_col(df: pd.DataFrame, includes: list[str], excludes: list[str] | None = None) -> str:
    excludes = excludes or []
    cols = list(df.columns)
    for c in cols:
        cc = clean_text(c)
        if all(clean_text(i) in cc for i in includes) and not any(clean_text(e) in cc for e in excludes):
            return c
    raise KeyError(f'列识别失败 includes={includes} excludes={excludes} 可选={cols}')


def pick_col_any(df: pd.DataFrame, candidates: list[tuple[list[str], list[str]]]) -> str:
    """按优先级识别列。candidates=[(必须包含关键词, 排除关键词), ...]"""
    last_error: Exception | None = None
    for includes, excludes in candidates:
        try:
            return pick_col(df, includes, excludes)
        except Exception as e:
            last_error = e
    raise KeyError(f'列识别失败，候选规则均未命中：{candidates}；最后错误：{last_error}')


def find_period_col(df: pd.DataFrame) -> str | None:
    """识别年份/期数列，用于从多期预测结果中筛出最后一年。"""
    strong_names = {'年份', '年度', '预测年份', '预测年度', '年序号', '第几年', '期数', '时间', '阶段'}
    for c in df.columns:
        cc = clean_text(c)
        if cc in strong_names:
            return c
    for c in df.columns:
        cc = clean_text(c)
        # 避免把“第5年末老人总数”“月实际需求”等指标列误判成年份列
        looks_like_period = re.search(r'(年份|年度|预测年|年序号|第几年|期数|阶段)', cc)
        looks_like_metric = any(k in cc for k in ['老人', '需求', '总数', '人数', '成本', '收入', '支出', '满意', '能力'])
        if looks_like_period and not looks_like_metric:
            return c
    return None


def filter_latest_period(df: pd.DataFrame, tag: str = '') -> pd.DataFrame:
    """若表中含年份/期数列，则只保留最后一期；否则原样返回。"""
    period_col = find_period_col(df)
    if period_col is None:
        return df.copy()
    tmp = df.copy()
    tmp['_period_num'] = tmp[period_col].apply(parse_number)
    valid = tmp.dropna(subset=['_period_num']).copy()
    if valid['_period_num'].nunique() <= 1:
        return tmp.drop(columns=['_period_num'])
    latest = valid['_period_num'].max()
    out = valid[valid['_period_num'] == latest].drop(columns=['_period_num'])
    print(f'INFO: {tag} 检测到多期数据，已按 {period_col}={latest:g} 筛选最后一期，行数 {len(df)} -> {len(out)}')
    return out


def collapse_metric_by_community(
    df: pd.DataFrame,
    community_col: str,
    value_col: str,
    value_name: str,
    agg: str = 'last'
) -> pd.DataFrame:
    """把重复小区记录压缩为一行，避免 10 个小区 × 多期 = 60 行。"""
    tmp = df[[community_col, value_col]].copy()
    tmp.columns = ['小区', value_name]
    tmp['小区'] = tmp['小区'].map(clean_text)
    tmp[value_name] = tmp[value_name].apply(parse_number)
    tmp = tmp.dropna(subset=[value_name])
    tmp = tmp[tmp['小区'] != ''].copy()

    if tmp['小区'].duplicated().any():
        before = len(tmp)
        if agg == 'sum':
            tmp = tmp.groupby('小区', as_index=False)[value_name].sum()
        elif agg == 'mean':
            tmp = tmp.groupby('小区', as_index=False)[value_name].mean()
        else:
            # 对“逐年预测表”最安全：取每个小区最后一条记录；若存在年份列，前面已筛最后一期
            tmp = tmp.groupby('小区', as_index=False)[value_name].last()
        print(f'INFO: {value_name} 存在重复小区记录，已按小区合并：{before} -> {len(tmp)}')

    return tmp


def parse_scale_name(s: str) -> str:
    t = clean_text(s)
    if '小' in t:
        return '小型'
    if '中' in t:
        return '中型'
    if '大' in t:
        return '大型'
    return t


def distance_satisfaction(d: float) -> float:
    if d <= 300:
        return 1.0
    if d <= 500:
        return 0.9
    if d <= 650:
        return 0.75
    if d <= 1000:
        return 0.6
    return 0.0


def response_satisfaction(u: float) -> float:
    if u <= 0.60:
        return 1.0
    if u <= 0.75:
        return 0.93
    if u <= 0.85:
        return 0.85
    if u <= 0.95:
        return 0.72
    if u <= 1.0:
        return 0.60
    return -1


@dataclass
class BestResult:
    key: tuple
    detail: dict


def safe_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        with open(path, 'a'):
            return path
    except PermissionError:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        return path.with_name(f'{path.stem}_{ts}{path.suffix}')



def clean_excel_cell(x: Any) -> Any:
    """去除 Excel XML 不允许的控制字符，避免生成的 xlsx 被 Excel 判定损坏。"""
    if isinstance(x, str):
        return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', x)
    return x


def clean_excel_df(df: pd.DataFrame) -> pd.DataFrame:
    """写入 Excel 前做轻量清洗：去掉 inf 和非法控制字符。"""
    out = df.copy()
    out = out.replace([np.inf, -np.inf], np.nan)
    for col in out.columns:
        if out[col].dtype == object:
            out[col] = out[col].map(clean_excel_cell)
    out.columns = [clean_excel_cell(c) for c in out.columns]
    return out


def write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, index: bool = False) -> None:
    """统一写表，确保 sheet 名合法且表格内容不会破坏 xlsx。"""
    invalid = r'[]:*?/\\'
    safe_name = ''.join('_' if ch in invalid else ch for ch in sheet_name)[:31]
    clean_excel_df(df).to_excel(writer, sheet_name=safe_name, index=index)


def write_validated_excel(output: Path, tables: list[tuple[str, pd.DataFrame, bool]]) -> Path:
    """先写临时 xlsx，验证可被 openpyxl 读取后再替换正式文件。

    这样可以避免运行中断、Excel 正打开文件、或写入一半时用户打开导致的损坏文件。
    """
    output = safe_output_path(output)
    tmp = output.with_name(f'{output.stem}__tmp_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

    with pd.ExcelWriter(tmp, engine='openpyxl') as writer:
        for sheet_name, df, index in tables:
            write_df(writer, df, sheet_name, index=index)

    if not zipfile.is_zipfile(tmp):
        raise RuntimeError(f'Excel临时文件不是有效xlsx压缩包: {tmp}')

    try:
        import openpyxl
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=False)
        _ = wb.sheetnames
        wb.close()
    except Exception as e:
        raise RuntimeError(f'Excel临时文件生成后校验失败: {tmp}, {e}') from e

    # 如果目标文件被 Excel 打开，Windows 下可能无法删除/覆盖；自动改用带时间戳的新文件名。
    try:
        if output.exists():
            output.unlink()
    except PermissionError:
        output = output.with_name(f'{output.stem}_{datetime.now().strftime("%Y%m%d_%H%M%S")}{output.suffix}')

    tmp.replace(output)
    print('INFO: Excel结果文件已通过有效性校验:', output)
    return output



def community_key(x: Any) -> str:
    """用于匹配“小区1 / 1小区 / 小区01 / A社区”等命名差异。"""
    s = clean_text(x)
    if s == '':
        return ''
    trans = str.maketrans('０１２３４５６７８９abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ',
                          '0123456789abcdefghijklmnopqrstuvwxyzabcdefghijklmnopqrstuvwxyz')
    s = s.translate(trans)
    s = s.replace('社区', '小区')
    s = re.sub(r'[\s_\-—－、,，:：;；（）()\[\]【】]', '', s)
    # 只要有“小区”和数字，就统一成“小区N”，避免“小区01”和“小区1”匹配不上。
    m = re.search(r'(\d+)', s)
    if '小区' in s and m:
        return f'小区{int(m.group(1))}'
    # 纯数字也按“小区N”处理，兼容距离矩阵行列名写成 1,2,...,10 的情况。
    if re.fullmatch(r'\d+(?:\.0)?', s):
        return f'小区{int(float(s))}'
    return s


def match_positions_to_communities(labels: list[Any], communities: list[str]) -> dict[str, int]:
    """把一组行/列标签匹配到题目中的10个小区，返回 {小区名: 位置}。"""
    key_to_comm: dict[str, str] = {}
    for c in communities:
        k = community_key(c)
        if k and k not in key_to_comm:
            key_to_comm[k] = c

    matched: dict[str, int] = {}
    used_pos: set[int] = set()

    # 1) 规范化完全匹配
    for pos, lab in enumerate(labels):
        k = community_key(lab)
        if k in key_to_comm and key_to_comm[k] not in matched:
            matched[key_to_comm[k]] = pos
            used_pos.add(pos)

    # 2) 唯一包含匹配，兼容“到小区1距离”“小区1(m)”这类标签
    for c in communities:
        if c in matched:
            continue
        ck = community_key(c)
        if not ck:
            continue
        cand = []
        for pos, lab in enumerate(labels):
            if pos in used_pos:
                continue
            lk = community_key(lab)
            if lk and (ck in lk or lk in ck):
                cand.append(pos)
        if len(cand) == 1:
            matched[c] = cand[0]
            used_pos.add(cand[0])

    return matched


def fill_symmetric_distance_matrix(dmat: pd.DataFrame) -> pd.DataFrame:
    """距离矩阵若只填了上三角/下三角，则用对称位置补齐。"""
    out = dmat.copy()
    for i in out.index:
        for j in out.columns:
            if i == j:
                out.loc[i, j] = 0.0
                continue
            a = out.loc[i, j]
            b = out.loc[j, i] if (j in out.index and i in out.columns) else np.nan
            if pd.isna(a) and not pd.isna(b):
                out.loc[i, j] = b
            elif not pd.isna(a) and pd.isna(b):
                out.loc[j, i] = a
    return out


def build_distance_matrix_from_labeled_df(df: pd.DataFrame, communities: list[str], sheet_name: str = '') -> pd.DataFrame | None:
    """处理已经有表头的距离矩阵：第一列为行小区，列名为列小区。"""
    if df.empty or len(df.columns) < 2:
        return None

    best: tuple[int, pd.DataFrame] | None = None
    for first_col in list(df.columns)[:min(5, len(df.columns))]:
        tmp = df.copy()
        row_labels = tmp[first_col].tolist()
        col_labels = list(tmp.columns)
        row_map = match_positions_to_communities(row_labels, communities)
        col_map = match_positions_to_communities(col_labels, communities)
        score = len(row_map) + len(col_map)
        if len(row_map) < max(3, len(communities) // 2) or len(col_map) < max(3, len(communities) // 2):
            continue

        mat = pd.DataFrame(np.nan, index=communities, columns=communities, dtype=float)
        for i in communities:
            if i not in row_map:
                continue
            rpos = row_map[i]
            for j in communities:
                if j not in col_map:
                    continue
                cpos = col_map[j]
                mat.loc[i, j] = parse_number(tmp.iloc[rpos, cpos])
        mat = fill_symmetric_distance_matrix(mat)
        if best is None or score > best[0]:
            best = (score, mat)

    return best[1] if best else None


def build_distance_matrix_from_raw_df(raw: pd.DataFrame, communities: list[str], sheet_name: str = '') -> pd.DataFrame | None:
    """处理带合并标题行/空白行的原始距离矩阵。"""
    raw = raw.dropna(how='all').dropna(axis=1, how='all').reset_index(drop=True)
    if raw.empty:
        return None

    best: tuple[int, pd.DataFrame, int, int] | None = None
    nrow, ncol = raw.shape
    # 在前15行、前6列内寻找“列小区标签行”和“行小区标签列”。
    for header_row in range(min(15, nrow)):
        header_vals = raw.iloc[header_row, :].tolist()
        col_map = match_positions_to_communities(header_vals, communities)
        if len(col_map) < max(3, len(communities) // 2):
            continue
        for label_col in range(min(6, ncol)):
            row_vals = raw.iloc[header_row + 1:, label_col].tolist()
            row_map_rel = match_positions_to_communities(row_vals, communities)
            if len(row_map_rel) < max(3, len(communities) // 2):
                continue
            row_map = {c: header_row + 1 + r for c, r in row_map_rel.items()}
            score = len(row_map) + len(col_map)
            mat = pd.DataFrame(np.nan, index=communities, columns=communities, dtype=float)
            for i in communities:
                if i not in row_map:
                    continue
                rpos = row_map[i]
                for j in communities:
                    if j not in col_map:
                        continue
                    cpos = col_map[j]
                    mat.loc[i, j] = parse_number(raw.iloc[rpos, cpos])
            mat = fill_symmetric_distance_matrix(mat)
            if best is None or score > best[0]:
                best = (score, mat, header_row, label_col)

    if best:
        _, mat, header_row, label_col = best
        print(f'INFO: 距离矩阵 sheet「{sheet_name}」已按原始表解析：列标签行={header_row + 1}，行标签列={label_col + 1}')
        return mat
    return None


def load_distance_matrix(path: Path, communities: list[str]) -> tuple[str, pd.DataFrame]:
    """稳健读取附件4距离矩阵，并输出缺失位置诊断。"""
    # 先用已有的自动表头逻辑读；失败再用 header=None 原始读取。
    parsed_sheets = read_workbook_sheets(path)
    candidates: list[tuple[str, pd.DataFrame]] = []
    for name, df in parsed_sheets.items():
        if '距离' in clean_text(name) or len(parsed_sheets) == 1:
            candidates.append((name, df))
    candidates += [(name, df) for name, df in parsed_sheets.items() if (name, df) not in candidates]

    best_name = ''
    best_mat: pd.DataFrame | None = None
    best_missing = 10**9
    for name, df in candidates:
        mat = build_distance_matrix_from_labeled_df(df, communities, sheet_name=name)
        if mat is None:
            continue
        missing = int(mat.isna().sum().sum())
        if missing < best_missing:
            best_name, best_mat, best_missing = name, mat, missing

    if best_mat is None or best_missing > 0:
        raw_sheets = pd.read_excel(path, sheet_name=None, header=None)
        for name, raw in raw_sheets.items():
            mat = build_distance_matrix_from_raw_df(raw, communities, sheet_name=name)
            if mat is None:
                continue
            missing = int(mat.isna().sum().sum())
            if missing < best_missing:
                best_name, best_mat, best_missing = name, mat, missing

    if best_mat is None:
        raise ValueError('无法解析附件4距离矩阵：没有找到同时包含小区行标签和列标签的矩阵。请检查附件4第一列和首行是否包含10个小区名称。')

    best_mat = fill_symmetric_distance_matrix(best_mat)
    for c in communities:
        best_mat.loc[c, c] = 0.0

    missing_pairs = []
    for i in communities:
        for j in communities:
            if i != j and pd.isna(best_mat.loc[i, j]):
                missing_pairs.append((i, j))

    if missing_pairs:
        print('ERROR: 距离矩阵仍有非对角缺失。前20个缺失位置如下：')
        for i, j in missing_pairs[:20]:
            print(f'  {i} -> {j}')
        print('INFO: 问题1结果中的小区名称:', communities)
        # 打印附件4可见行列标签，方便人工对照。
        for name, df in parsed_sheets.items():
            print(f'INFO: 附件4 sheet「{name}」列名预览:', list(df.columns)[:15])
            if len(df.columns) > 0:
                print(f'INFO: 附件4 sheet「{name}」第一列预览:', df.iloc[:15, 0].map(clean_text).tolist())
        raise ValueError(f'距离矩阵存在非对角缺失值：共 {len(missing_pairs)} 处。通常是附件4的小区名称与问题1输出的小区名称不一致，或附件4只有部分矩阵。')

    print(f'INFO: 距离矩阵读取成功，使用sheet「{best_name}」，规模={best_mat.shape[0]}×{best_mat.shape[1]}')
    return best_name, best_mat


# =========================
# 问题2：允许小区需求拆分后的 MILP 求解部分
# =========================

# 服务响应满意度分段规则。边界按附件5理解：低利用率优先取得更高分。
RESPONSE_TIERS = [
    {'tier': 'u≤0.60',        'lower': 0.00, 'upper': 0.60, 's2': 1.00},
    {'tier': '0.60<u≤0.75',  'lower': 0.60, 'upper': 0.75, 's2': 0.93},
    {'tier': '0.75<u≤0.85',  'lower': 0.75, 'upper': 0.85, 's2': 0.85},
    {'tier': '0.85<u≤0.95',  'lower': 0.85, 'upper': 0.95, 's2': 0.72},
    {'tier': '0.95<u≤1.00',  'lower': 0.95, 'upper': 1.00, 's2': 0.60},
]

# MILP 求解时间上限；如果电脑较慢，可以调大到 300 或 600。
MILP_TIME_LIMIT_SEC = 600
MILP_REL_GAP = 1e-6
NUM_TOL = 1e-6


def require_scipy_milp():
    """导入 scipy MILP 求解器。"""
    try:
        from scipy.optimize import milp, LinearConstraint, Bounds
        from scipy.sparse import lil_matrix, csr_matrix
        return milp, LinearConstraint, Bounds, lil_matrix, csr_matrix
    except Exception as e:
        raise ImportError(
            '本版 problem2_split_milp.py 使用 scipy.optimize.milp 求解“可拆分需求”的混合整数线性规划。'
            '请先在当前虚拟环境安装或升级 scipy：\n'
            '  pip install -U scipy\n'
            '或：\n'
            '  conda install scipy\n'
            f'原始错误：{e}'
        ) from e


def make_sparse_constraint(coeff: np.ndarray, lb: float = -np.inf, ub: float = np.inf):
    """把一行系数向量转为 scipy LinearConstraint。"""
    _, LinearConstraint, _, _, csr_matrix = require_scipy_milp()
    return LinearConstraint(csr_matrix(coeff.reshape(1, -1)), np.array([lb], dtype=float), np.array([ub], dtype=float))


def solve_split_demand_milp(
    communities: list[str],
    P: dict[str, float],
    Qd: dict[str, float],
    dmat: pd.DataFrame,
    params: dict[str, dict[str, float]],
    weights: dict[str, float],
) -> dict[str, Any]:
    """允许一个小区需求拆分给多个服务站的快速 MILP 求解。

    决策变量：
    1. y[j,s] ∈ {0,1}：是否在小区 j 建设规模 s 的服务站；
    2. x[i,j,s] ≥ 0：小区 i 的日均需求中有多少分配给小区 j 的规模 s 服务站；
    3. Umax：所有服务站中的最大利用率。

    约束：
    - 每个候选小区最多建设一个服务站；
    - 总建设成本不超过预算；
    - 每个小区分配出去的需求不超过其日均需求；
    - 只有建站后才能接收需求，且接收量不超过该规模服务能力；
    - 只允许距离不超过 SERVICE_RADIUS_M 的小区-站点对分配；
    - Umax 不小于任一服务站利用率。

    分层优化（更强调覆盖率与满意度，同时保持较快求解）：
    1. 最大化等价覆盖人口；
    2. 在覆盖人口最优下，最大化需求覆盖量；
    3. 在前两者最优下，最大化距离-价格满意度代理值；
    4. 在前三者最优下，最小化建设成本；
    5. 在前四者最优下，最大化总安装服务能力；
    6. 在前五者最优下，最小化最大利用率，以间接提高响应满意度。

    说明：响应满意度 S2 是利用率的阶梯函数，直接线性化会显著增加整数变量。
    本版在求解中使用“最大总安装能力 + 最小最大利用率”作为响应质量代理，
    最终结果表与图仍按附件5的利用率分段规则精确复核 S2 和综合满意度。
    """
    milp, LinearConstraint, Bounds, lil_matrix, csr_matrix = require_scipy_milp()

    scale_names = [s for s in ['小型', '中型', '大型'] if s in params]
    if not scale_names:
        raise ValueError('没有识别到小型/中型/大型服务站参数，请检查附件3。')

    feasible_pairs = [
        (i, j)
        for i in communities
        for j in communities
        if float(dmat.loc[i, j]) <= SERVICE_RADIUS_M and Qd.get(i, 0) > 0
    ]
    if not feasible_pairs:
        raise RuntimeError('没有任何小区-服务站候选对满足服务半径约束。')

    v_idx: dict[tuple[str, str, int], int] = {}
    x_idx: dict[tuple[str, str, str, int], int] = {}
    idx = 0

    for j in communities:
        for s in scale_names:
            v_idx[(j, s, 0)] = idx
            idx += 1

    for i, j in feasible_pairs:
        for s in scale_names:
            x_idx[(i, j, s, 0)] = idx
            idx += 1

    umax_idx = idx
    idx += 1
    n_var = idx

    lb = np.zeros(n_var, dtype=float)
    ub = np.full(n_var, np.inf, dtype=float)
    integrality = np.zeros(n_var, dtype=int)

    for k in v_idx.values():
        ub[k] = 1.0
        integrality[k] = 1
    ub[umax_idx] = 1.0

    rows: list[dict[int, float]] = []
    lows: list[float] = []
    ups: list[float] = []

    def add_row(coefs: dict[int, float], low: float = -np.inf, up: float = np.inf) -> None:
        rows.append(coefs)
        lows.append(low)
        ups.append(up)

    # 每个候选小区最多建一个服务站。
    for j in communities:
        coefs = {v_idx[(j, s, 0)]: 1.0 for s in scale_names}
        add_row(coefs, -np.inf, 1.0)

    # 总建设预算。
    budget_coefs = {}
    for (j, s, _t), k in v_idx.items():
        budget_coefs[k] = float(params[s]['build'])
    add_row(budget_coefs, -np.inf, float(BUDGET_YUAN))

    # 每个小区最多分配自身全部日均需求。
    for i in communities:
        coefs = {}
        for (ii, j, s, _t), k in x_idx.items():
            if ii == i:
                coefs[k] = 1.0
        add_row(coefs, -np.inf, float(Qd[i]))

    # 服务站容量上界：sum_i x[i,j,s] <= cap_s * y[j,s]
    # 最大利用率：sum_i x[i,j,s] / cap_s <= Umax
    for j in communities:
        for s in scale_names:
            cap = float(params[s]['cap'])
            if cap <= 0:
                raise ValueError(f'服务站规模 {s} 的日服务能力无效：{cap}')
            y = v_idx[(j, s, 0)]
            load_terms = {}
            for (i2, j2, s2, _t), k in x_idx.items():
                if j2 == j and s2 == s:
                    load_terms[k] = 1.0

            cap_row = dict(load_terms)
            cap_row[y] = cap_row.get(y, 0.0) - cap
            add_row(cap_row, -np.inf, 0.0)

            umax_row = {k: val / cap for k, val in load_terms.items()}
            umax_row[umax_idx] = umax_row.get(umax_idx, 0.0) - 1.0
            add_row(umax_row, -np.inf, 0.0)

    A = lil_matrix((len(rows), n_var), dtype=float)
    for r, coefs in enumerate(rows):
        for c, val in coefs.items():
            A[r, c] = val
    base_constraint = LinearConstraint(A.tocsr(), np.array(lows, dtype=float), np.array(ups, dtype=float))

    coverage_coeff = np.zeros(n_var, dtype=float)
    sat_coeff = np.zeros(n_var, dtype=float)
    demand_coeff = np.zeros(n_var, dtype=float)
    build_coeff = np.zeros(n_var, dtype=float)
    capacity_coeff = np.zeros(n_var, dtype=float)
    umax_coeff = np.zeros(n_var, dtype=float)
    umax_coeff[umax_idx] = 1.0

    for (j, s, _t), k in v_idx.items():
        build_coeff[k] = float(params[s]['build'])
        capacity_coeff[k] = float(params[s]['cap'])

    for (i, j, s, _t), k in x_idx.items():
        q = float(Qd[i])
        pop_per_daily = float(P[i]) / q if q > 0 else 0.0
        s1 = distance_satisfaction(float(dmat.loc[i, j]))
        s3 = 1.0
        sij_proxy = weights['distance'] * s1 + weights['price'] * s3
        coverage_coeff[k] = pop_per_daily
        sat_coeff[k] = pop_per_daily * sij_proxy
        demand_coeff[k] = 30.0

    bounds = Bounds(lb, ub)

    def run_stage(
        coeff: np.ndarray,
        maximize: bool,
        extra_constraints: list[Any],
        stage_name: str,
    ):
        c = -coeff if maximize else coeff.copy()
        res = milp(
            c=c,
            integrality=integrality,
            bounds=bounds,
            constraints=[base_constraint] + extra_constraints,
            options={'time_limit': MILP_TIME_LIMIT_SEC, 'mip_rel_gap': 1e-6, 'disp': False}
        )
        if not res.success:
            raise RuntimeError(f'MILP阶段「{stage_name}」求解失败：status={res.status}, message={res.message}')
        val = float(coeff @ res.x)
        print(f'INFO: MILP阶段「{stage_name}」完成，目标值={val:.6g}')
        return res, val

    constraints_extra: list[Any] = []

    res1, opt_cov = run_stage(coverage_coeff, True, constraints_extra, '1-最大覆盖人口')
    tol_cov = max(1e-4, abs(opt_cov) * 1e-6)
    constraints_extra.append(make_sparse_constraint(coverage_coeff, lb=opt_cov - tol_cov, ub=np.inf))

    res2, opt_dem = run_stage(demand_coeff, True, constraints_extra, '2-最大需求覆盖')
    tol_dem = max(1e-4, abs(opt_dem) * 1e-6)
    constraints_extra.append(make_sparse_constraint(demand_coeff, lb=opt_dem - tol_dem, ub=np.inf))

    res3, opt_sat = run_stage(sat_coeff, True, constraints_extra, '3-最大距离价格满意度')
    tol_sat = max(1e-4, abs(opt_sat) * 1e-6)
    constraints_extra.append(make_sparse_constraint(sat_coeff, lb=opt_sat - tol_sat, ub=np.inf))

    res4, opt_build = run_stage(build_coeff, False, constraints_extra, '4-最小建设成本')
    tol_build = max(1e-2, abs(opt_build) * 1e-6)
    constraints_extra.append(make_sparse_constraint(build_coeff, lb=-np.inf, ub=opt_build + tol_build))

    res5, opt_cap = run_stage(capacity_coeff, True, constraints_extra, '5-最大总安装能力')
    tol_cap = max(1e-4, abs(opt_cap) * 1e-6)
    constraints_extra.append(make_sparse_constraint(capacity_coeff, lb=opt_cap - tol_cap, ub=np.inf))

    res6, opt_umax = run_stage(umax_coeff, False, constraints_extra, '6-最小最大利用率')
    x = res6.x

    return {
        'x': x,
        'v_idx': v_idx,
        'x_idx': x_idx,
        'umax_idx': umax_idx,
        'coverage_coeff': coverage_coeff,
        'sat_coeff': sat_coeff,
        'demand_coeff': demand_coeff,
        'build_coeff': build_coeff,
        'capacity_coeff': capacity_coeff,
        'umax_coeff': umax_coeff,
        'opt_cov': float(coverage_coeff @ x),
        'opt_sat_num_proxy': float(sat_coeff @ x),
        'opt_dem_month': float(demand_coeff @ x),
        'opt_build': float(build_coeff @ x),
        'opt_capacity': float(capacity_coeff @ x),
        'opt_umax': float(umax_coeff @ x),
        'scale_names': scale_names,
    }

def extract_split_solution(
    milp_sol: dict[str, Any],
    communities: list[str],
    P: dict[str, float],
    Qd: dict[str, float],
    Qm: dict[str, float],
    dmat: pd.DataFrame,
    params: dict[str, dict[str, float]],
    weights: dict[str, float],
    demand_type: pd.DataFrame,
    price_df: pd.DataFrame,
) -> dict[str, pd.DataFrame | dict | float]:
    """把 MILP 变量转为结果表，并按实际利用率复核 S2。"""
    x = milp_sol['x']
    v_idx = milp_sol['v_idx']
    x_idx = milp_sol['x_idx']

    selected: dict[str, dict[str, Any]] = {}
    for (j, s, _t_idx), k in v_idx.items():
        if x[k] > 0.5:
            selected[j] = {
                '规模': s,
                '日服务能力': float(params[s]['cap']),
                '建设成本': float(params[s]['build']),
                '日固定管理成本': float(params[s]['fixed_day']),
            }

    # 先统计流量和站点负载。
    raw_flows: list[tuple[str, str, str, float]] = []
    station_load = {j: 0.0 for j in selected}
    community_served_daily = {i: 0.0 for i in communities}

    for (i, j, s, _t_idx), k in x_idx.items():
        val = float(x[k])
        if val <= NUM_TOL:
            continue
        if j not in selected:
            continue
        raw_flows.append((i, j, s, val))
        station_load[j] += val
        community_served_daily[i] += val

    station_s2 = {}
    station_util = {}
    for j, info in selected.items():
        cap = info['日服务能力']
        util = station_load.get(j, 0.0) / cap if cap > 0 else 0.0
        station_util[j] = util
        station_s2[j] = response_satisfaction(util)

    # 再按实际站点利用率计算每条分流的满意度。
    flow_rows = []
    community_sat_num = {i: 0.0 for i in communities}
    for i, j, s, val in raw_flows:
        s1 = distance_satisfaction(float(dmat.loc[i, j]))
        s2 = station_s2[j]
        s3 = 1.0
        sij = weights['distance'] * s1 + weights['response'] * s2 + weights['price'] * s3
        fraction_i_j = val / float(Qd[i]) if Qd[i] > 0 else 0.0
        pop_equiv = float(P[i]) * fraction_i_j
        community_sat_num[i] += pop_equiv * sij

        flow_rows.append({
            '小区': i,
            '服务站': j,
            '服务站规模': s,
            '日分配需求': val,
            '月分配需求': val * 30.0,
            '占本小区需求比例': fraction_i_j,
            '等价覆盖老人数量': pop_equiv,
            '距离': float(dmat.loc[i, j]),
            '距离满意度S1': s1,
            '服务站实际利用率': station_util[j],
            '响应满意度S2': s2,
            '价格满意度S3': s3,
            '综合满意度S': sij,
        })

    flow_df = pd.DataFrame(flow_rows)

    station_rows = []
    for j, info in selected.items():
        ld = station_load.get(j, 0.0)
        cap = info['日服务能力']
        util = station_util[j]
        station_rows.append({
            '站点': j,
            '规模': info['规模'],
            '建设成本': info['建设成本'],
            '日服务能力': cap,
            '日固定管理成本': info['日固定管理成本'],
            '年固定管理成本': info['日固定管理成本'] * 365,
            '日实际服务量': ld,
            '利用率': util,
            '响应满意度S2': station_s2[j],
        })
    station_df = pd.DataFrame(station_rows)

    alloc_rows = []
    for i in communities:
        served = community_served_daily[i]
        q = float(Qd[i])
        frac = min(1.0, served / q) if q > 0 else 0.0
        avg_sat_i = community_sat_num[i] / (float(P[i]) * frac) if frac > NUM_TOL and P[i] > 0 else 0.0
        sub = flow_df[flow_df['小区'] == i] if not flow_df.empty else pd.DataFrame()
        if sub.empty:
            assign_txt = '未覆盖'
        else:
            parts = [
                f"{r['服务站']}({r['占本小区需求比例']:.1%})"
                for _, r in sub.sort_values('占本小区需求比例', ascending=False).iterrows()
            ]
            assign_txt = '、'.join(parts)

        alloc_rows.append({
            '小区': i,
            '第5年末老人总数': float(P[i]),
            '月实际需求': float(Qm[i]),
            '日均需求': q,
            '日已分配需求': served,
            '日未满足需求': max(0.0, q - served),
            '需求满足比例': frac,
            '等价覆盖老人数量': float(P[i]) * frac,
            '分配服务站及比例': assign_txt,
            '综合满意度': avg_sat_i,
            '是否完全覆盖': int(frac >= 1 - 1e-5),
            '是否部分覆盖': int((frac > 1e-5) and (frac < 1 - 1e-5)),
            '是否未覆盖': int(frac <= 1e-5),
        })
    alloc_df = pd.DataFrame(alloc_rows)

    cover_rows = []
    for j in selected:
        sub = flow_df[flow_df['服务站'] == j] if not flow_df.empty else pd.DataFrame()
        cover_rows.append({
            '站点': j,
            '覆盖小区及比例': '、'.join(
                f"{r['小区']}({r['占本小区需求比例']:.1%})"
                for _, r in sub.sort_values(['小区', '占本小区需求比例']).iterrows()
            ) if not sub.empty else '',
            '等价覆盖老人总数': sub['等价覆盖老人数量'].sum() if not sub.empty else 0.0,
            '覆盖月需求': sub['月分配需求'].sum() if not sub.empty else 0.0,
            '覆盖日需求': sub['日分配需求'].sum() if not sub.empty else 0.0,
            '利用率': station_load[j] / selected[j]['日服务能力'] if selected[j]['日服务能力'] > 0 else 0.0,
            '响应满意度S2': station_s2[j],
        })
    cover_df = pd.DataFrame(cover_rows)

    # 按“小区-服务项目”精确分摊到站点，而不是使用全局平均单价。
    if not flow_df.empty:
        fraction_df = flow_df[['小区', '服务站', '占本小区需求比例']].copy()
        service_alloc = demand_type.merge(fraction_df, on='小区', how='inner')
        service_alloc['站点月服务量'] = service_alloc['月需求'] * service_alloc['占本小区需求比例']
        service_alloc = service_alloc.merge(price_df, on='服务项目', how='left')
        service_alloc['收入'] = service_alloc['站点月服务量'] * service_alloc['单价']
        service_alloc['直接成本'] = service_alloc['站点月服务量'] * service_alloc['直接支出']
    else:
        service_alloc = pd.DataFrame(columns=['服务站', '收入', '直接成本'])

    profit_rows = []
    for j, info in selected.items():
        sub = service_alloc[service_alloc['服务站'] == j] if not service_alloc.empty else pd.DataFrame()
        annual_rev = float(sub['收入'].sum() * 12) if not sub.empty else 0.0
        annual_dc = float(sub['直接成本'].sum() * 12) if not sub.empty else 0.0
        fixed = info['日固定管理成本'] * 365
        build = info['建设成本']
        dep = build / 20.0
        profit_rows.append({
            '站点': j,
            '年度收入': annual_rev,
            '年直接支出': annual_dc,
            '年固定管理成本': fixed,
            '建设成本': build,
            '年折旧': dep,
            '运营利润': annual_rev - annual_dc - fixed,
            '含折旧利润': annual_rev - annual_dc - fixed - dep,
        })
    profit_df = pd.DataFrame(profit_rows)

    total_pop = sum(float(P[i]) for i in communities)
    total_month = sum(float(Qm[i]) for i in communities)
    covered_pop = float(alloc_df['等价覆盖老人数量'].sum())
    covered_month = float(alloc_df['日已分配需求'].sum() * 30.0)
    sat_num = sum(community_sat_num[i] for i in communities)
    avg_sat = sat_num / max(covered_pop, 1e-9)

    overall = pd.DataFrame([
        ('服务覆盖率(等价人口)', covered_pop / total_pop if total_pop > 0 else 0.0),
        ('人口加权平均满意度', avg_sat),
        ('等价覆盖老人数量', covered_pop),
        ('总老人数量', total_pop),
        ('覆盖月需求', covered_month),
        ('总月需求', total_month),
        ('需求覆盖率', covered_month / total_month if total_month > 0 else 0.0),
        ('完全覆盖小区数', int(alloc_df['是否完全覆盖'].sum())),
        ('部分覆盖小区数', int(alloc_df['是否部分覆盖'].sum())),
        ('未覆盖小区数', int(alloc_df['是否未覆盖'].sum())),
        ('总建设成本', float(station_df['建设成本'].sum()) if not station_df.empty else 0.0),
        ('总日服务能力', float(station_df['日服务能力'].sum()) if not station_df.empty else 0.0),
        ('总日需求', total_month / 30.0 if total_month > 0 else 0.0),
        ('总日已分配需求', float(alloc_df['日已分配需求'].sum())),
        ('最大利用率', float(station_df['利用率'].max()) if not station_df.empty else 0.0),
        ('站点数量', len(selected)),
        ('小型数量', sum(1 for v in selected.values() if v['规模'] == '小型')),
        ('中型数量', sum(1 for v in selected.values() if v['规模'] == '中型')),
        ('大型数量', sum(1 for v in selected.values() if v['规模'] == '大型')),
    ], columns=['指标', '数值'])

    return {
        'built': selected,
        'station_df': station_df,
        'flow_df': flow_df,
        'alloc_df': alloc_df,
        'cover_df': cover_df,
        'profit_df': profit_df,
        'overall': overall,
        'service_alloc': service_alloc,
        'covered_pop': covered_pop,
        'covered_month': covered_month,
        'avg_sat': avg_sat,
    }


def save_split_charts(
    alloc_df: pd.DataFrame,
    station_df: pd.DataFrame,
    flow_df: pd.DataFrame,
    dmat: pd.DataFrame,
    communities: list[str],
    built: dict[str, dict[str, Any]],
) -> None:
    charts = Path('charts_problem2_split')
    charts.mkdir(exist_ok=True)
    if not HAS_MPL:
        print('WARNING: matplotlib 不可用，已跳过图表生成。')
        return

    def add_bar_labels(ax, fmt: str = '{:.3f}', ypad_ratio: float = 0.01) -> None:
        ymin, ymax = ax.get_ylim()
        span = max(ymax - ymin, 1e-9)
        for p in ax.patches:
            h = p.get_height()
            x = p.get_x() + p.get_width() / 2
            ax.text(x, h + span * ypad_ratio, fmt.format(h), ha='center', va='bottom', fontsize=8)

    uncovered = alloc_df.loc[alloc_df['是否未覆盖'] == 1, '小区'].tolist()
    uncovered_text = '、'.join(uncovered) if uncovered else '无'

    # 图2：各小区满意度柱状图（带坐标轴、数值标注、未覆盖小区说明）
    fig, ax = plt.subplots(figsize=(10, 6))
    x = np.arange(len(alloc_df))
    vals = alloc_df['综合满意度'].to_numpy(dtype=float)
    colors = ['#d62728' if int(u) == 1 else '#1f77b4' for u in alloc_df['是否未覆盖']]
    bars = ax.bar(x, vals, color=colors, width=0.55)
    ax.set_xticks(x)
    ax.set_xticklabels(alloc_df['小区'])
    ax.set_xlabel('小区')
    ax.set_ylabel('综合满意度')
    ax.set_title('各小区满意度（可拆分需求）')
    ax.set_ylim(0, max(1.0, float(vals.max()) * 1.12 if len(vals) else 1.0))
    ax.grid(axis='y', linestyle='--', alpha=0.35)
    add_bar_labels(ax, fmt='{:.3f}', ypad_ratio=0.01)
    ax.text(0.98, 0.98, f'未覆盖小区：{uncovered_text}', transform=ax.transAxes,
            ha='right', va='top', fontsize=9,
            bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.85, edgecolor='0.7'))
    plt.tight_layout()
    plt.savefig(charts / '02_各小区满意度柱状图.png', dpi=180)
    plt.close(fig)

    # 图3：各小区需求满足比例（带坐标轴、数值标注）
    fig, ax = plt.subplots(figsize=(10, 6))
    vals = alloc_df['需求满足比例'].to_numpy(dtype=float)
    colors = ['#d62728' if int(u) == 1 else '#2ca02c' if int(p) == 0 else '#ff7f0e'
              for u, p in zip(alloc_df['是否未覆盖'], alloc_df['是否完全覆盖'])]
    ax.bar(x, vals, color=colors, width=0.55)
    ax.set_xticks(x)
    ax.set_xticklabels(alloc_df['小区'])
    ax.set_xlabel('小区')
    ax.set_ylabel('需求满足比例')
    ax.set_title('各小区需求满足比例（可拆分需求）')
    ax.set_ylim(0, 1.12)
    ax.grid(axis='y', linestyle='--', alpha=0.35)
    add_bar_labels(ax, fmt='{:.1%}', ypad_ratio=0.01)
    ax.text(0.98, 0.98, f'未覆盖小区：{uncovered_text}', transform=ax.transAxes,
            ha='right', va='top', fontsize=9,
            bbox=dict(boxstyle='round,pad=0.25', facecolor='white', alpha=0.85, edgecolor='0.7'))
    plt.tight_layout()
    plt.savefig(charts / '03_各小区需求满足比例.png', dpi=180)
    plt.close(fig)

    # 图4：各服务站利用率柱状图（加数值标注）
    if not station_df.empty:
        fig, ax = plt.subplots(figsize=(9, 6))
        sx = np.arange(len(station_df))
        vals = station_df['利用率'].to_numpy(dtype=float)
        ax.bar(sx, vals, width=0.55)
        ax.set_xticks(sx)
        ax.set_xticklabels(station_df['站点'])
        ax.set_xlabel('服务站')
        ax.set_ylabel('利用率')
        ax.set_title('各服务站利用率（可拆分需求）')
        ax.set_ylim(0, max(1.05, float(vals.max()) * 1.12 if len(vals) else 1.0))
        for y in [0.60, 0.75, 0.85, 0.95, 1.00]:
            ax.axhline(y, linestyle='--', linewidth=0.8, alpha=0.25)
        ax.grid(axis='y', linestyle='--', alpha=0.35)
        add_bar_labels(ax, fmt='{:.3f}', ypad_ratio=0.01)
        plt.tight_layout()
        plt.savefig(charts / '04_各服务站利用率柱状图.png', dpi=180)
        plt.close(fig)

    def classical_mds_positions(distance_df: pd.DataFrame, labels: list[str]) -> dict[str, tuple[float, float]]:
        D = distance_df.reindex(index=labels, columns=labels).astype(float).to_numpy()
        D = np.nan_to_num(D, nan=0.0)
        D = (D + D.T) / 2.0
        np.fill_diagonal(D, 0.0)
        n = D.shape[0]
        J = np.eye(n) - np.ones((n, n)) / n
        B = -0.5 * J @ (D ** 2) @ J
        eigvals, eigvecs = np.linalg.eigh(B)
        idx = np.argsort(eigvals)[::-1]
        eigvals = eigvals[idx]
        eigvecs = eigvecs[:, idx]
        coords = np.zeros((n, 2), dtype=float)
        positive_dims = [k for k, v in enumerate(eigvals) if v > 1e-9][:2]
        for out_dim, eig_idx in enumerate(positive_dims):
            coords[:, out_dim] = eigvecs[:, eig_idx] * np.sqrt(eigvals[eig_idx])
        if len(positive_dims) < 2 or np.ptp(coords[:, 1]) < 1e-9:
            coords[:, 1] = np.linspace(-200.0, 200.0, n)
        # 只做平移，不做缩放，尽量保留“米”量级，便于粗略判断是否接近1000m。
        coords[:, 0] = coords[:, 0] - coords[:, 0].mean()
        coords[:, 1] = coords[:, 1] - coords[:, 1].mean()
        return {lab: (float(coords[i, 0]), float(coords[i, 1])) for i, lab in enumerate(labels)}

    pos = classical_mds_positions(dmat, communities)
    fig, ax = plt.subplots(figsize=(10, 7))

    # 拆分流量线：线宽表示比例，颜色区分是否超过1000m（理论上不应超过）。
    if not flow_df.empty:
        for _, r in flow_df.iterrows():
            i, j = r['小区'], r['服务站']
            xi, yi = pos[i]
            xj, yj = pos[j]
            width = 0.8 + 4.0 * min(1.0, float(r['占本小区需求比例']))
            actual_d = float(r['距离'])
            color = '#d62728' if actual_d > SERVICE_RADIUS_M + 1e-9 else '#7f7f7f'
            ax.plot([xi, xj], [yi, yj], linewidth=width, alpha=0.45, color=color, zorder=1)

    uncovered_set = set(uncovered)
    covered_normal = [c for c in communities if c not in built and c not in uncovered_set]

    if covered_normal:
        xs = [pos[c][0] for c in covered_normal]
        ys = [pos[c][1] for c in covered_normal]
        ax.scatter(xs, ys, s=45, marker='o', color='#ff7f0e', label='已覆盖普通小区', zorder=2)
        for c in covered_normal:
            ax.text(pos[c][0], pos[c][1] + 30, c, ha='center', va='bottom', fontsize=8)

    if uncovered_set:
        xs = [pos[c][0] for c in uncovered_set]
        ys = [pos[c][1] for c in uncovered_set]
        ax.scatter(xs, ys, s=55, marker='o', color='#d62728', label='未覆盖小区', zorder=2)
        for c in uncovered_set:
            ax.text(pos[c][0], pos[c][1] + 30, c, ha='center', va='bottom', fontsize=8, color='#d62728', fontweight='bold')

    if built:
        xs = [pos[c][0] for c in built]
        ys = [pos[c][1] for c in built]
        ax.scatter(xs, ys, s=110, marker='*', color='#1f77b4', label='建站小区', zorder=3)
        for c in built:
            ax.text(pos[c][0], pos[c][1] + 45, f'{c}\n{built[c]["规模"]}', ha='center', va='bottom', fontsize=8, fontweight='bold')

    all_x = [v[0] for v in pos.values()]
    all_y = [v[1] for v in pos.values()]
    xmin, xmax = min(all_x), max(all_x)
    ymin, ymax = min(all_y), max(all_y)
    xspan = max(xmax - xmin, 1.0)
    yspan = max(ymax - ymin, 1.0)
    ax.set_xlim(xmin - 0.12 * xspan, xmax + 0.28 * xspan)
    ax.set_ylim(ymin - 0.12 * yspan, ymax + 0.15 * yspan)
    ax.set_aspect('equal', adjustable='box')
    ax.grid(True, linestyle='--', alpha=0.25)
    ax.set_xlabel('MDS近似坐标 X（米）')
    ax.set_ylabel('MDS近似坐标 Y（米）')
    ax.set_title('最优站点覆盖关系图')

    # 1000米比例尺，便于粗略判断。    
    bar_len = 1000.0
    bar_x0 = xmin + 0.68 * xspan
    bar_y0 = ymin - 0.05 * yspan
    ax.plot([bar_x0, bar_x0 + bar_len], [bar_y0, bar_y0], color='black', linewidth=2.0)
    ax.plot([bar_x0, bar_x0], [bar_y0 - 0.02 * yspan, bar_y0 + 0.02 * yspan], color='black', linewidth=1.5)
    ax.plot([bar_x0 + bar_len, bar_x0 + bar_len], [bar_y0 - 0.02 * yspan, bar_y0 + 0.02 * yspan], color='black', linewidth=1.5)
    ax.text(bar_x0 + bar_len / 2, bar_y0 + 0.04 * yspan, '1000 m 参考尺度', ha='center', va='bottom', fontsize=9)

    ax.legend(loc='upper right', fontsize=8, frameon=False)
    plt.tight_layout()
    plt.savefig(charts / '01_最优站点覆盖关系图_可拆分.png', dpi=180)
    plt.close(fig)

def main() -> None:
    root = Path('.')
    p1 = root / 'B_problem1_results.xlsx'
    a2 = root / 'data/附件2：服务需求数据.xlsx'
    a3 = root / 'data/附件3：服务站建设与运营成本.xlsx'
    a4 = root / 'data/附件4：小区间距离矩阵.xlsx'
    a5 = root / 'data/附件5：满意度评分规则.xlsx'

    p1s = read_workbook_sheets(p1)
    s02n, s02 = choose_sheet_by_keywords(p1s, ['老人', '逐小区'])
    s08n, s08 = choose_sheet_by_keywords(p1s, ['实际需求', '汇总'])
    s07n, s07 = choose_sheet_by_keywords(p1s, ['实际需求', '类型'])

    s02_latest = filter_latest_period(s02, tag='老人逐小区表')
    s08_latest = filter_latest_period(s08, tag='实际需求汇总表')
    s07_latest = filter_latest_period(s07, tag='实际需求类型表')

    comm_col = pick_col(s02_latest, ['小区'])
    p_col = pick_col_any(s02_latest, [
        (['第5', '老人', '总'], []),
        (['5', '老人', '总'], []),
        (['年末', '老人', '总'], []),
        (['老人', '总'], []),
        (['总'], [])
    ])
    df_pop = collapse_metric_by_community(
        s02_latest,
        community_col=comm_col,
        value_col=p_col,
        value_name='第5年末老人总数',
        agg='last'
    )

    comm2 = pick_col(s08_latest, ['小区'])
    q_col = pick_col_any(s08_latest, [
        (['第5', '实际', '合计'], []),
        (['实际', '合计'], []),
        (['月', '实际'], []),
        (['实际'], ['率'])
    ])
    df_dem = collapse_metric_by_community(
        s08_latest,
        community_col=comm2,
        value_col=q_col,
        value_name='月实际需求',
        agg='last'
    )

    base = df_pop.merge(df_dem, on='小区', how='inner')
    base['日均需求q'] = base['月实际需求'] / 30.0
    if len(base) != 10:
        raise ValueError(
            f'小区数量应为10，当前={len(base)}，唯一小区数={base["小区"].nunique()}。'
            f'老人表小区数={df_pop["小区"].nunique()}，需求表小区数={df_dem["小区"].nunique()}。'
            '请检查 B_problem1_results.xlsx 中小区名称是否一致，或是否混入了非小区汇总行。'
        )

    a3s = read_workbook_sheets(a3)
    s3n, s3 = choose_sheet_by_keywords(a3s, ['规模'])
    size_col = pick_col(s3, ['规模'])
    cost_col = pick_col_any(s3, [
        (['建设', '成本'], []),
        (['建设'], []),
        (['成本'], ['固定', '运营', '管理'])
    ])
    fixed_col = pick_col_any(s3, [
        (['日', '固定'], []),
        (['固定', '成本'], []),
        (['运营', '成本'], []),
        (['管理', '成本'], []),
        (['日'], ['能力', '服务能力'])
    ])
    cap_col = pick_col_any(s3, [
        (['日', '服务', '能力'], []),
        (['服务', '能力'], []),
        (['最大', '服务', '人次'], []),
        (['服务', '人次'], []),
        (['人次'], ['成本', '固定', '运营', '管理']),
        (['能力'], ['成本', '固定', '运营', '管理'])
    ])
    scale_df = s3[[size_col, cost_col, fixed_col, cap_col]].copy()
    scale_df.columns = ['规模', '建设成本原', '日固定成本', '日服务能力']
    scale_df['规模'] = scale_df['规模'].map(parse_scale_name)
    scale_df = scale_df[scale_df['规模'].isin(['小型', '中型', '大型'])].copy()
    scale_df['建设成本原'] = scale_df['建设成本原'].apply(parse_number)
    scale_df['日固定成本'] = scale_df['日固定成本'].apply(parse_number)
    scale_df['日服务能力'] = scale_df['日服务能力'].apply(parse_number)
    max_build = scale_df['建设成本原'].max()
    scale_df['建设成本'] = scale_df['建设成本原'] * (10000 if max_build < 1000 else 1)
    scale_df['年固定成本'] = scale_df['日固定成本'] * 365

    communities = base['小区'].tolist()
    s4n, dmat = load_distance_matrix(a4, communities)

    a5s = read_workbook_sheets(a5)
    weights = parse_satisfaction_weights(a5s)

    # 小区-服务项目月需求，用于利润精确分摊。
    p1_comm = pick_col(s07_latest, ['小区'])
    p1_type = pick_col(s07_latest, ['服务'])
    p1_qty = pick_col_any(s07_latest, [
        (['第5', '实际'], []),
        (['月', '实际'], []),
        (['实际'], ['率'])
    ])
    demand_type = s07_latest[[p1_comm, p1_type, p1_qty]].copy()
    demand_type.columns = ['小区', '服务项目', '月需求']
    demand_type['小区'] = demand_type['小区'].map(clean_text)
    demand_type['服务项目'] = demand_type['服务项目'].map(clean_text)
    demand_type['月需求'] = demand_type['月需求'].apply(parse_number)
    demand_type = demand_type.dropna(subset=['月需求'])
    demand_type = demand_type[(demand_type['小区'] != '') & (demand_type['服务项目'] != '')]
    demand_type = demand_type.groupby(['小区', '服务项目'], as_index=False)['月需求'].sum()

    a2s = read_workbook_sheets(a2)
    s2pn, s2p = choose_sheet_by_keywords(a2s, ['营收'])
    c_service = pick_col(s2p, ['服务项目'])
    c_price = pick_col(s2p, ['营收'])
    c_direct = pick_col(s2p, ['支出'])
    price_df = s2p[[c_service, c_price, c_direct]].copy()
    price_df.columns = ['服务项目', '单价', '直接支出']
    price_df['服务项目'] = price_df['服务项目'].map(clean_text)
    price_df['单价'] = price_df['单价'].apply(parse_number)
    price_df['直接支出'] = price_df['直接支出'].apply(parse_number)

    params = {
        r['规模']: {
            'build': float(r['建设成本']),
            'fixed_day': float(r['日固定成本']),
            'cap': float(r['日服务能力'])
        }
        for _, r in scale_df.iterrows()
    }

    P = dict(zip(base['小区'], base['第5年末老人总数']))
    Qd = dict(zip(base['小区'], base['日均需求q']))
    Qm = dict(zip(base['小区'], base['月实际需求']))

    print('开始求解允许拆分需求的快速 MILP 模型...')
    print('候选建站点数:', len(communities), '候选点:', communities)
    print('预算上限:', BUDGET_YUAN, '服务半径:', SERVICE_RADIUS_M)
    milp_sol = solve_split_demand_milp(communities, P, Qd, dmat, params, weights)
    result = extract_split_solution(milp_sol, communities, P, Qd, Qm, dmat, params, weights, demand_type, price_df)

    station_df = result['station_df']
    flow_df = result['flow_df']
    alloc_df = result['alloc_df']
    cover_df = result['cover_df']
    profit_df = result['profit_df']
    overall = result['overall']
    service_alloc = result['service_alloc']

    data_check_df = pd.DataFrame([
        {'文件': 'B_problem1_results.xlsx', 'sheet总数': len(p1s), '使用sheet': f'{s02n},{s07n},{s08n}'},
        {'文件': '附件2', 'sheet总数': len(a2s), '使用sheet': s2pn},
        {'文件': '附件3', 'sheet总数': len(a3s), '使用sheet': s3n},
        {'文件': '附件4', 'sheet总数': '自动解析', '使用sheet': s4n},
        {'文件': '附件5', 'sheet总数': len(a5s), '使用sheet': '满意度评分规则'},
        {'文件': '求解模型', 'sheet总数': '-', '使用sheet': '快速MILP：候选小区建站 + 小区需求可拆分'}
    ])

    model_note = pd.DataFrame([
        {'项目': '建站位置假设', '说明': '不允许任意地点建站，仅允许在10个小区候选点中选择建站。'},
        {'项目': '需求分配假设', '说明': '允许一个小区的日均服务需求按比例拆分给多个服务半径内的服务站。'},
        {'项目': '覆盖率定义', '说明': '若某小区只满足部分需求，则按“已满足需求/总需求”的比例折算等价覆盖老人数量。'},
        {'项目': '价格满意度', '说明': '问题2不调整服务价格，沿用附件2基准价格，因此价格满意度S3取1.00。'},
        {'项目': '优化方法', '说明': '使用简化混合整数线性规划MILP，分层优化：覆盖人口→距离价格满意度→需求覆盖→建设成本→最大利用率；最终按实际利用率复核响应满意度S2。'},
    ])

    output = write_validated_excel(Path('B_problem2_split_results.xlsx'), [
        ('01_数据读取检查', data_check_df, False),
        ('02_问题1输入数据', base, False),
        ('03_服务站规模参数', scale_df[['规模', '建设成本', '日固定成本', '日服务能力']], False),
        ('04_距离矩阵', dmat.reset_index(), False),
        ('05_最优选址规模方案', station_df, False),
        ('06_小区需求满足汇总', alloc_df, False),
        ('07_小区到站点分流明细', flow_df, False),
        ('08_服务站覆盖明细', cover_df, False),
        ('09_年度利润测算', profit_df, False),
        ('10_服务项目收入分摊', service_alloc, False),
        ('11_总体指标', overall, False),
        ('12_模型说明', model_note, False),
    ])

    save_split_charts(alloc_df, station_df, flow_df, dmat, communities, result['built'])

    print('=== 最终最优方案摘要（允许拆分需求） ===')
    print('站点数量:', len(result['built']))
    print('方案:', {k: v['规模'] for k, v in result['built'].items()})
    print('服务覆盖率(等价人口):', result['covered_pop'] / sum(P.values()))
    print('平均满意度:', result['avg_sat'])
    print('总建设成本:', station_df['建设成本'].sum() if not station_df.empty else 0)
    print('需求覆盖率:', result['covered_month'] / sum(Qm.values()))
    print('完全覆盖小区数:', int(alloc_df['是否完全覆盖'].sum()))
    print('部分覆盖小区数:', int(alloc_df['是否部分覆盖'].sum()))
    print('未覆盖小区数:', int(alloc_df['是否未覆盖'].sum()))
    print('结果文件:', output)
    print('图表目录: charts_problem2_split/')


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print('ERROR:', e)
        sys.exit(1)
